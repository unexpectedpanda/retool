#!/usr/bin/env python

# An easier way to visualize parent/clone relationships in a Redump 1G1R dat
# Used to verify the output of Retool.

# There isn't much error checking in here, it's rough and only intended to
# work for a limited use case.

import html
import os
import re
import sys
import textwrap

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import cell as Cell

def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    print(font.bold + '\nClonerel 0.14' + font.end)
    print('-------------')

    if len(sys.argv) == 1:
        print(
            textwrap.fill(
            'Creates parent/clone visualizations from a Retool-generated 1G1R'
            + ' Redump dat.'
            )
        )

        command = ''

        if 'clonerel.py' in sys.argv[0]:
            command = 'python '

        print('\nUSAGE: ' + font.bold + command + os.path.basename(sys.argv[0]) + font.end + ' <input dat>')
        sys.exit()

    # Super basic validation
    if sys.argv[1].endswith('.dat'):
        input_file_name = os.path.abspath(sys.argv[1])

        if not os.path.exists(input_file_name):
            print(textwrap.TextWrapper(width=80, subsequent_indent='  ').fill(font.red + '* Input file "' + font.bold + input_file_name + font.end + font.red + '" does not exist.' + font.end))
        else:
            print('* Reading file...')
            with open(input_file_name, 'r') as input_file_read:
                checkdat = input_file_read.read()
                soup = BeautifulSoup(checkdat, "lxml-xml")

            print('* Finding parents...')
            parent_list = {}
            orphan_list = []

            # Restrict scope to just clones
            clone_soup = soup.find_all('game', {'cloneof':re.compile('.*')})

            # Generate the parent list from those clones
            for item in clone_soup:
                parent_title = re.search('cloneof=".*?"', str(item)).group()[9:-1]
                if parent_title not in parent_list:
                    parent_list[parent_title] = []

            # Now add the clones to each of those parents
            for item in clone_soup:
                parent_title = re.search('cloneof=".*?"', str(item)).group()[9:-1]
                clone_title = re.search('name=".*?"', str(item)).group()[6:-1]

                if parent_title in parent_list:
                    parent_list[parent_title].append(clone_title)

            # Now add titles with no clones
            all_soup = soup.find_all('game', {'name':re.compile('.*')})
            for item in all_soup:
                title = re.search('name=".*?"', str(item)).group()[6:-1]
                has_clone = bool(re.search('cloneof=".*?"', str(item)))
                if has_clone == False and title not in parent_list:
                    orphan_list.append(title)

            parent_list_sorted = {}

            # Do some formatting
            sort_list = sorted(parent_list.keys())

            for item in sort_list:
                parent_list_sorted[item] = parent_list[item]


            # Create a parents + orphans list for easier title comparison
            # in the Excel file
            parent_orphan_list = []

            for item in parent_list_sorted.keys():
                parent_orphan_list.append(item)

            for item in orphan_list:
                parent_orphan_list.append(item)

            # Create an Excel spreadsheet
            print('* Creating Excel file...')
            wb = Workbook()
            ws = wb.active

            ws.title = 'Parents'

            col = 'A'
            row = 2

            # Add the header
            ws.merge_cells('A1:B1')
            ws['A1'] = 'Parents with clones'
            ws['C1'] = 'Parents with orphans'
            ws['D1'] = 'Orphans'
            ws['A1'].font = Font(bold = True, color = 'ffffffff', size = '12')
            ws['A1'].fill = PatternFill("solid", fgColor="ff808080")
            ws['A1'].alignment = Alignment(vertical = 'center')
            ws['C1'].font = Font(bold = True, color = 'ffffffff', size = '12')
            ws['C1'].fill = PatternFill("solid", fgColor="ff808080")
            ws['C1'].alignment = Alignment(vertical = 'center')
            ws['D1'].font = Font(bold = True, color = 'ffffffff', size = '12')
            ws['D1'].fill = PatternFill("solid", fgColor="ff808080")
            ws['D1'].alignment = Alignment(vertical = 'center')

            ws.row_dimensions[1].height = 20
            ws.freeze_panes = ws['A2']

            # Populate parents that have clones
            for item in parent_list_sorted:
                ws.merge_cells(col + str(row) + ':' + chr(ord(col) + 1) + str(row))
                ws[col + str(row)] = html.unescape(item)
                ws[col + str(row)].font = Font(bold = True, size = 12)

                for i, subitem in enumerate(sorted(parent_list_sorted[item])):
                    row += 1
                    if i < len(parent_list_sorted[item]) - 1:
                        ws[chr(ord(col)) + str(row)].alignment = Alignment(horizontal = 'right')
                        ws[chr(ord(col)) + str(row)] = '├'
                    else:
                        ws[chr(ord(col)) + str(row)].alignment = Alignment(horizontal = 'right')
                        ws[chr(ord(col)) + str(row)] = '└'

                    ws[chr(ord(col) + 1) + str(row)] = html.unescape(subitem)
                    if i == len(parent_list_sorted[item]) - 1:
                        ws[chr(ord(col)) + str(row + 1)] = ''
                        row += 2

            # Populate all parents
            col = 'C'
            row = 2
            for item in sorted(parent_orphan_list):
                ws[col + str(row)] = html.unescape(item)
                row += 1

            # Populate orphans
            col = 'D'
            row = 2
            for item in sorted(orphan_list):
                ws[col + str(row)] = html.unescape(item)
                row += 1

            # Adjust column widths
            dims = {}

            for col in ws.rows:
                for cell in col:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
            for col, value in dims.items():
                if col != 1:
                    ws.column_dimensions[Cell.get_column_letter(col)].width = value + 5


            ws.column_dimensions[Cell.get_column_letter(1)].width = 5

            # Write the file to disk
            py_file = sys.argv[0]
            path_name = os.path.dirname(py_file)
            file_path = os.path.join(path_name, os.path.basename(input_file_name)[:-3] + 'xlsx')
            print('* Outputting to "' + file_path + font.end + '...')
            wb.save(file_path)
            print('\nDone!')

############### Classes and methods ###############
# Console text formatting
class font:
    purple = '\033[95m'
    cyan = '\033[96m'
    darkcyan = '\033[36m'
    blue = '\033[94m'
    green = '\033[92m'
    yellow = '\033[93m'
    white = '\033[37m'
    red = '\033[91m'
    bold = '\033[1m'
    underline = '\033[4m'
    end = '\033[0m'
    blink = '\033[5m'

if __name__ == '__main__':
    main()